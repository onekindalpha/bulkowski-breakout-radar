#!/usr/bin/env python3
"""
make_buy_report_v6.py  (BUY_NOW first + colored Excel + PRIORITY ranking + optional PDF)

- Adds --pdf (matplotlib table; no reportlab needed)
- Produces CSV always, XLSX when --xlsx, PDF when --pdf
"""
from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd

KST = ZoneInfo("Asia/Seoul")


def newest_report_path() -> Path:
    stamped = sorted(Path(".").glob("report_v2_*_KST.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
    if stamped:
        return stamped[0]
    p = Path("report_v2.csv")
    if not p.exists():
        raise FileNotFoundError("No report found: report_v2.csv or report_v2_*_KST.csv")
    return p


def _fmt_bool(x) -> str:
    try:
        return "Y" if bool(x) else "N"
    except Exception:
        return "N"


def _to_float(x):
    try:
        return float(x)
    except Exception:
        return None


def _clamp(x, lo, hi):
    return max(lo, min(hi, x))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--top", type=int, default=20, help="rows in buy_report (after ranking)")
    ap.add_argument("--tickers", default="", help="comma tickers; overrides --top")
    ap.add_argument("--xlsx", action="store_true", help="write a colored Excel file too")
    ap.add_argument("--pdf", action="store_true", help="write a PDF table too (matplotlib)")
    ap.add_argument("--buy-chase-pct", type=float, default=2.0)
    ap.add_argument("--buy-min-room-pct", type=float, default=1.0)
    ap.add_argument("--buy-rsi-min", type=float, default=45.0)
    ap.add_argument("--buy-rsi-max", type=float, default=75.0)
    ap.add_argument("--require-intraday", action="store_true")
    ap.add_argument("--no-intraday", action="store_true")
    args = ap.parse_args()

    rp = newest_report_path()
    df = pd.read_csv(rp, comment="#")
    df.columns = [c.strip() for c in df.columns]

    want_cols = [
        "ticker","grade","score","price","gap_pct","rsi14",
        "daily_break_level","daily_breakout","daily_retest",
        "breakout_confirmed_close","breakout_volume_confirmed","breakout_confirmed_intraday",
        "weekly_r1","room_to_weekly_r1_pct","px_vs_sma50","px_vs_sma200",
    ]
    cols = [c for c in want_cols if c in df.columns]
    out = df[cols].copy()

    if args.tickers.strip():
        keep = {t.strip().upper() for t in args.tickers.split(",") if t.strip()}
        out = out[out["ticker"].astype(str).str.upper().isin(keep)].copy()

    has_intraday_col = "breakout_confirmed_intraday" in out.columns
    if args.no_intraday:
        require_intraday = False
    elif args.require_intraday:
        require_intraday = True
    else:
        require_intraday = has_intraday_col

    def build_reason(r):
        parts = []
        if "grade" in out.columns and pd.notna(r.get("grade")):
            parts.append(f"Grade={r.get('grade')}")
        if "score" in out.columns and pd.notna(r.get("score")):
            parts.append(f"Score={r.get('score')}")
        if "breakout_confirmed_intraday" in out.columns:
            parts.append(f"IntradayHold={_fmt_bool(r.get('breakout_confirmed_intraday'))}")
        if "breakout_confirmed_close" in out.columns:
            parts.append(f"CloseBreak={_fmt_bool(r.get('breakout_confirmed_close'))}")
        if "breakout_volume_confirmed" in out.columns:
            parts.append(f"VolConfirm={_fmt_bool(r.get('breakout_volume_confirmed'))}")
        if "daily_retest" in out.columns:
            parts.append(f"Retest={_fmt_bool(r.get('daily_retest'))}")
        if "rsi14" in out.columns and pd.notna(r.get("rsi14")):
            parts.append(f"RSI={float(r.get('rsi14')):.1f}")
        if "room_to_weekly_r1_pct" in out.columns and pd.notna(r.get("room_to_weekly_r1_pct")):
            parts.append(f"RoomR1={float(r.get('room_to_weekly_r1_pct')):.1f}%")
        if "gap_pct" in out.columns and pd.notna(r.get("gap_pct")):
            parts.append(f"Gap={float(r.get('gap_pct')):.2f}%")
        if "daily_break_level" in out.columns and pd.notna(r.get("daily_break_level")):
            parts.append(f"BreakLvl={float(r.get('daily_break_level')):.2f}")
        return " | ".join(parts)

    out["reason"] = out.apply(build_reason, axis=1)

    def breakout_pct(r):
        price = _to_float(r.get("price"))
        lvl = _to_float(r.get("daily_break_level"))
        if price is None or lvl is None or lvl == 0:
            return None
        return (price / lvl - 1.0) * 100.0

    out["breakout_pct_vs_level"] = out.apply(breakout_pct, axis=1)
    out["chase_flag"] = out["breakout_pct_vs_level"].apply(
        lambda x: "CHASE" if (x is not None and pd.notna(x) and x > args.buy_chase_pct) else ""
    )

    def buy_eval(r):
        fail = []
        if "breakout_confirmed_close" in out.columns:
            if not bool(r.get("breakout_confirmed_close")):
                fail.append("CloseBreak=N")
        else:
            fail.append("CloseBreak=missing")

        if "breakout_volume_confirmed" in out.columns:
            if not bool(r.get("breakout_volume_confirmed")):
                fail.append("VolConfirm=N")
        else:
            fail.append("VolConfirm=missing")

        if require_intraday:
            if not has_intraday_col:
                fail.append("IntradayHold=missing")
            else:
                if not bool(r.get("breakout_confirmed_intraday")):
                    fail.append("IntradayHold=N")

        price = _to_float(r.get("price"))
        lvl = _to_float(r.get("daily_break_level"))
        rsi = _to_float(r.get("rsi14"))
        room = _to_float(r.get("room_to_weekly_r1_pct"))

        if price is None or lvl is None:
            fail.append("Price/Level=missing")
        else:
            if price < lvl:
                fail.append("BelowBreakLvl")
            if price > lvl * (1.0 + args.buy_chase_pct/100.0):
                fail.append(f"Chase>{args.buy_chase_pct:.1f}%")

        if room is None:
            fail.append("RoomR1=missing")
        else:
            if room < args.buy_min_room_pct:
                fail.append(f"RoomR1<{args.buy_min_room_pct:.1f}%")

        if rsi is None:
            fail.append("RSI=missing")
        else:
            if rsi < args.buy_rsi_min:
                fail.append(f"RSI<{args.buy_rsi_min:.0f}")
            if rsi > args.buy_rsi_max:
                fail.append(f"RSI>{args.buy_rsi_max:.0f}")

        buy_now = (len(fail) == 0)
        return ("Y" if buy_now else "N"), ("OK" if buy_now else ";".join(fail))

    out[["BUY_NOW","BUY_NOW_REASON"]] = out.apply(lambda r: pd.Series(buy_eval(r)), axis=1)

    def is_near(r):
        return (bool(r.get("breakout_confirmed_close")) and bool(r.get("breakout_volume_confirmed")) and r.get("BUY_NOW") == "N")
    out["near_buy"] = out.apply(is_near, axis=1)

    def priority_score(r):
        score = 0.0
        if bool(r.get("breakout_confirmed_close")): score += 20
        if bool(r.get("breakout_volume_confirmed")): score += 20
        if "breakout_confirmed_intraday" in out.columns and bool(r.get("breakout_confirmed_intraday")): score += 15
        if bool(r.get("daily_retest")): score += 10

        room = _to_float(r.get("room_to_weekly_r1_pct"))
        if room is not None:
            score += _clamp(room, 0, 10) * 2.0

        pct = _to_float(r.get("breakout_pct_vs_level"))
        if pct is not None:
            if pct < 0:
                score += -5
            elif pct <= 1:
                score += 20 - pct*5
            elif pct <= 2:
                score += 15 - (pct-1)*10
            else:
                score += 5 - _clamp((pct-2)*5, 0, 20)

        rsi = _to_float(r.get("rsi14"))
        if rsi is not None:
            score += _clamp(10 - abs(rsi-60)/3, 0, 10)

        sma50 = _to_float(r.get("px_vs_sma50"))
        if sma50 is not None:
            score += _clamp(sma50, -5, 10) * 0.5

        return round(score, 2)

    out["PRIORITY_SCORE"] = out.apply(priority_score, axis=1)

    out["BUY_NOW_sort"] = out["BUY_NOW"].map({"Y":1,"N":0}).fillna(0).astype(int)
    out = out.sort_values(["BUY_NOW_sort","PRIORITY_SCORE","score"], ascending=[False,False,False]).drop(columns=["BUY_NOW_sort"])

    if not args.tickers.strip():
        out = out.head(max(1, args.top)).copy()

    out["PRIORITY_RANK"] = range(1, len(out)+1)

    front = ["BUY_NOW","BUY_NOW_REASON","PRIORITY_RANK","PRIORITY_SCORE","near_buy","chase_flag",
             "ticker","price","daily_break_level","breakout_pct_vs_level","reason"]
    new_cols = []
    for c in front:
        if c in out.columns and c not in new_cols:
            new_cols.append(c)
    for c in out.columns:
        if c not in new_cols:
            new_cols.append(c)
    out = out[new_cols].copy()

    ts = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
    csv_name = f"buy_report_{ts}_KST.csv"
    out.to_csv(csv_name, index=False)
    print(f"Saved: {csv_name}")

    if args.xlsx:
        try:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.formatting.rule import FormulaRule

            xlsx_name = f"buy_report_{ts}_KST.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "buy_report"

            for r in dataframe_to_rows(out, index=False, header=True):
                ws.append(r)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            header = [c.value for c in ws[1]]
            buy_col = header.index("BUY_NOW") + 1 if "BUY_NOW" in header else 1
            near_col = header.index("near_buy") + 1 if "near_buy" in header else None
            chase_col = header.index("chase_flag") + 1 if "chase_flag" in header else None

            fill_buy = PatternFill("solid", fgColor="C6EFCE")
            fill_near = PatternFill("solid", fgColor="FFF2CC")
            fill_chase = PatternFill("solid", fgColor="F8CBAD")
            fill_n = PatternFill("solid", fgColor="F2F2F2")
            font_buy = Font(bold=True, color="006100")
            align = Alignment(vertical="center")

            last_row = ws.max_row
            last_col = ws.max_column
            rng = f"A2:{ws.cell(row=last_row, column=last_col).coordinate}"
            buy_letter = ws.cell(row=1, column=buy_col).column_letter

            ws.conditional_formatting.add(rng, FormulaRule(formula=[f'${buy_letter}2="Y"'], fill=fill_buy, font=font_buy))
            if near_col is not None:
                near_letter = ws.cell(row=1, column=near_col).column_letter
                ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(${buy_letter}2="N",${near_letter}2=TRUE)'], fill=fill_near))
            if chase_col is not None:
                chase_letter = ws.cell(row=1, column=chase_col).column_letter
                ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(${buy_letter}2="N",${chase_letter}2="CHASE")'], fill=fill_chase))
            ws.conditional_formatting.add(rng, FormulaRule(formula=[f'${buy_letter}2="N"'], fill=fill_n))

            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align

            for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
                for cell in row:
                    cell.alignment = align

            for j, name in enumerate(header, 1):
                ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = max(10, min(42, len(str(name)) + 2))

            wb.save(xlsx_name)
            print(f"Saved: {xlsx_name}")
        except Exception as e:
            print(f"[warn] XLSX not created: {type(e).__name__}: {e}")

    if args.pdf:
        try:
            import math
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_pdf import PdfPages

            pdf_name = f"buy_report_{ts}_KST.pdf"
            cols_pdf = [c for c in ["BUY_NOW","PRIORITY_RANK","ticker","price","daily_break_level",
                                    "breakout_pct_vs_level","PRIORITY_SCORE","near_buy","chase_flag","BUY_NOW_REASON"]
                        if c in out.columns]
            view = out[cols_pdf].copy()
            rows_per_page = 22
            pages = math.ceil(len(view) / rows_per_page)

            with PdfPages(pdf_name) as pdf:
                for pi in range(pages):
                    chunk = view.iloc[pi*rows_per_page:(pi+1)*rows_per_page]
                    fig, ax = plt.subplots(figsize=(11.69, 8.27))
                    ax.axis("off")
                    title = f"Buy Report (KST {datetime.now(KST).strftime('%Y-%m-%d %H:%M:%S')})  Source={rp.name}  Page {pi+1}/{pages}"
                    ax.set_title(title, fontsize=10, pad=10)
                    table = ax.table(cellText=chunk.values.tolist(), colLabels=list(chunk.columns), loc="center")
                    table.auto_set_font_size(False)
                    table.set_fontsize(8)
                    table.scale(1, 1.3)
                    pdf.savefig(fig, bbox_inches="tight")
                    plt.close(fig)
            print(f"Saved: {pdf_name}")
        except Exception as e:
            print(f"[warn] PDF not created: {type(e).__name__}: {e}")

if __name__ == "__main__":
    main()
