#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
국장 엑셀 워크북 보정 스크립트
- 모든 시트 보존
- KRX(pykrx) 최신 종가 + 상장주식수 반영
- 액면분할/병합 등으로 상장주식수가 바뀐 종목의 per-share 값 정규화
- 가격 연동 컬럼 재계산
- 타이밍자동체크 시트의 현재가도 최신 종가 기준으로 맞춤

주의:
- pykrx는 장중 실시간 틱이 아니라 최신 거래일 종가/시총/상장주식수를 가져오는 데 쓰는 게 안전함.
- 장중 실시간이 꼭 필요하면 별도 quote collector를 만들어 "현재가_실시간" 컬럼으로만 덮는 걸 권장.
"""

from __future__ import annotations

import argparse
import math
from datetime import date, timedelta
from pathlib import Path
from typing import Optional, Dict, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

try:
    from pykrx import stock
except Exception as e:
    raise SystemExit(f"pykrx import 실패: {e}")


MAIN_REQUIRED = {"종목코드", "종목명", "현재가"}
PRICE_REQUIRED = {"종목코드", "종목명", "현재가"}


def to_float(x):
    if x is None or x == "":
        return None
    if isinstance(x, str):
        s = x.replace(",", "").strip()
        if s in {"", "-", "nan", "None"}:
            return None
        x = s
    try:
        v = float(x)
        if math.isnan(v):
            return None
        return v
    except Exception:
        return None


def normalize_code(x) -> Optional[str]:
    if x is None:
        return None
    s = "".join(ch for ch in str(x) if ch.isdigit())
    if not s:
        return None
    return s.zfill(6)[-6:]


def safe_div(a, b):
    a = to_float(a)
    b = to_float(b)
    if a is None or b in (None, 0):
        return None
    return a / b


def find_header_row(ws, required: set[str], max_rows: int = 10) -> Optional[int]:
    for r in range(1, min(max_rows, ws.max_row) + 1):
        vals = {ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if required.issubset(vals):
            return r
    return None


def header_map(ws, header_row: int) -> Dict[str, int]:
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            out[v.strip()] = c
    return out


def cell(ws, row: int, hmap: Dict[str, int], name: str):
    col = hmap.get(name)
    return ws.cell(row, col) if col else None


def getv(ws, row: int, hmap: Dict[str, int], name: str):
    c = cell(ws, row, hmap, name)
    return None if c is None else c.value


def setv(ws, row: int, hmap: Dict[str, int], name: str, value):
    c = cell(ws, row, hmap, name)
    if c is not None:
        c.value = value


def latest_ohlcv_and_shares(code: str, asof: Optional[str] = None, lookback_days: int = 180) -> Tuple[str, float, int]:
    """
    반환: (기준일, 최신종가, 최신상장주식수)
    """
    end_dt = date.today() if asof is None else date.fromisoformat(asof)
    start_dt = end_dt - timedelta(days=lookback_days)

    ohlcv = stock.get_market_ohlcv_by_date(start_dt.strftime("%Y%m%d"), end_dt.strftime("%Y%m%d"), code)
    cap = stock.get_market_cap_by_date(start_dt.strftime("%Y%m%d"), end_dt.strftime("%Y%m%d"), code)

    if ohlcv is None or ohlcv.empty:
        raise RuntimeError(f"OHLCV 조회 실패: {code}")
    if cap is None or cap.empty:
        raise RuntimeError(f"시총/상장주식수 조회 실패: {code}")

    close = float(ohlcv["종가"].astype(float).iloc[-1])
    shares = int(float(cap["상장주식수"].astype(float).iloc[-1]))
    latest_date = str(ohlcv.index[-1].date()) if hasattr(ohlcv.index[-1], "date") else str(ohlcv.index[-1])

    return latest_date, close, shares


def recalc_graham_values(ws, row: int, hmap: Dict[str, int]):
    eps = to_float(getv(ws, row, hmap, "EPS(FY2025)"))
    if eps is None:
        return

    # 적정PER은 성장률 기반이라 분할과 무관. 기존 값 유지.
    for horizon in ["1년", "3년", "5년"]:
        per_col = f"그레이엄적정PER({horizon})"
        iv_col = f"그레이엄내재가치({horizon})"
        if per_col in hmap and iv_col in hmap:
            perv = to_float(getv(ws, row, hmap, per_col))
            if perv is not None:
                setv(ws, row, hmap, iv_col, eps * perv)

    # 선택 기준도 다시 맞춤
    use_basis = getv(ws, row, hmap, "그레이엄사용기준")
    if isinstance(use_basis, str):
        use_basis = use_basis.strip()
    iv_sel = None
    per_sel = None
    if use_basis in {"1년", "3년", "5년"}:
        per_sel = to_float(getv(ws, row, hmap, f"그레이엄적정PER({use_basis})"))
        iv_sel = to_float(getv(ws, row, hmap, f"그레이엄내재가치({use_basis})"))
    if per_sel is not None:
        setv(ws, row, hmap, "그레이엄적정PER(선택)", per_sel)
    if iv_sel is not None:
        setv(ws, row, hmap, "그레이엄내재가치(선택)", iv_sel)


def recalc_price_dependent(ws, row: int, hmap: Dict[str, int]):
    price = to_float(getv(ws, row, hmap, "현재가"))
    eps = to_float(getv(ws, row, hmap, "EPS(FY2025)"))
    dps = to_float(getv(ws, row, hmap, "현금배당금(FY2025 DPS)"))
    netcash_l = to_float(getv(ws, row, hmap, "주당순현금(린치식)"))
    netcash_c = to_float(getv(ws, row, hmap, "주당순현금(보수형)"))
    fcf_ps = to_float(getv(ws, row, hmap, "주당잉여현금흐름"))

    if price not in (None, 0) and eps not in (None, 0):
        if netcash_l is not None:
            setv(ws, row, hmap, "순현금차감PER(린치식)", (price - netcash_l) / eps)
        if netcash_c is not None:
            setv(ws, row, hmap, "순현금차감PER(보수형)", (price - netcash_c) / eps)

    # Graham 괴리율
    if price not in (None, 0):
        for horizon in ["1년", "3년", "5년"]:
            iv_col = f"그레이엄내재가치({horizon})"
            disc_col = f"그레이엄괴리율({horizon},%)"
            iv = to_float(getv(ws, row, hmap, iv_col))
            if iv is not None:
                setv(ws, row, hmap, disc_col, (iv / price - 1) * 100)

        iv_sel = to_float(getv(ws, row, hmap, "그레이엄내재가치(선택)"))
        if iv_sel is not None:
            setv(ws, row, hmap, "그레이엄괴리율(선택,%)", (iv_sel / price - 1) * 100)

    # 배당수익률
    if price not in (None, 0) and dps is not None:
        setv(ws, row, hmap, "배당수익률(%)", (dps / price) * 100)

    # 잉여현금흐름수익률
    if price not in (None, 0) and fcf_ps is not None:
        setv(ws, row, hmap, "잉여현금흐름수익률(%)", (fcf_ps / price) * 100)

    # 배당감안이익성장률, 배당감안점수
    div_y = to_float(getv(ws, row, hmap, "배당수익률(%)"))
    ncper_l = to_float(getv(ws, row, hmap, "순현금차감PER(린치식)"))

    growth_map = {
        "1년": "연간이익증가율(1년,%)",
        "3년": "연간이익증가율(3년CAGR,%)",
        "5년": "연간이익증가율(5년CAGR,%)",
    }
    score_map = {
        "1년": "배당감안이익성장률(1년)",
        "3년": "배당감안이익성장률(3년)",
        "5년": "배당감안이익성장률(5년)",
    }

    chosen = {}
    for k, gcol in growth_map.items():
        g = to_float(getv(ws, row, hmap, gcol))
        v = None
        if g is not None and div_y is not None and ncper_l not in (None, 0):
            v = (g + div_y) / ncper_l
        setv(ws, row, hmap, score_map[k], v)
        chosen[k] = v

    basis = getv(ws, row, hmap, "배당감안점수기준")
    basis = str(basis).strip() if basis is not None else "3년"
    if basis not in chosen:
        basis = "3년"
    setv(ws, row, hmap, "배당감안점수", chosen.get(basis))

    # 린치PER배수 = 순현금차감PER / 성장률(선택)
    lynch_basis = getv(ws, row, hmap, "린치기준")
    lynch_basis = str(lynch_basis).strip() if lynch_basis is not None else "3년"
    gcol = {
        "1년": "연간이익증가율(1년,%)",
        "3년": "연간이익증가율(3년CAGR,%)",
        "5년": "연간이익증가율(5년CAGR,%)",
    }.get(lynch_basis, "연간이익증가율(3년CAGR,%)")
    g = to_float(getv(ws, row, hmap, gcol))
    lmult = None
    if ncper_l not in (None, 0) and g not in (None, 0):
        lmult = ncper_l / g
    setv(ws, row, hmap, "린치PER배수", lmult)

    # 단순 판정 갱신
    if lmult is not None:
        if lmult <= 0.5:
            judge = "매우 저평가"
        elif lmult <= 1.0:
            judge = "유망"
        elif lmult <= 1.5:
            judge = "보통"
        else:
            judge = "고평가"
        setv(ws, row, hmap, "린치PER판정", judge)

    # 종합판정은 기존 하드필터 존중
    hard_ok = (
        (to_float(getv(ws, row, hmap, "주당순현금(린치식)")) or 0) > 0 and
        (to_float(getv(ws, row, hmap, "주당잉여현금흐름")) or 0) > 0 and
        (to_float(getv(ws, row, hmap, "순현금차감PER(린치식)")) or 0) > 0
    )
    lmult = to_float(getv(ws, row, hmap, "린치PER배수"))
    graham_gap = to_float(getv(ws, row, hmap, "그레이엄괴리율(선택,%)"))

    if not hard_ok:
        final = "제외"
    elif lmult is None or graham_gap is None:
        final = getv(ws, row, hmap, "종합판정")
    else:
        if lmult <= 1.0 and graham_gap >= 0:
            final = "매우 유망"
        elif lmult <= 1.5 and graham_gap >= -20:
            final = "유망"
        else:
            final = "보류"
    setv(ws, row, hmap, "종합판정", final)


def normalize_per_share_values(ws, row: int, hmap: Dict[str, int], latest_shares: int):
    """
    핵심:
    - 액면분할/병합이 있으면 발행주식수 차이로 factor 계산
    - absolute totals가 있으면 per-share는 totals / latest_shares 로 다시 계산
    - totals가 없고 per-share만 있는 항목(EPS, DPS, 주당FCF)은 factor로 환산
    """
    stored_shares = to_float(getv(ws, row, hmap, "발행주식수"))
    if stored_shares in (None, 0):
        setv(ws, row, hmap, "발행주식수", latest_shares)
        return

    factor = latest_shares / stored_shares
    setv(ws, row, hmap, "발행주식수", latest_shares)

    # absolute totals 기반 재계산
    cash = to_float(getv(ws, row, hmap, "현금및현금성자산")) or 0
    sec = to_float(getv(ws, row, hmap, "유가증권성자산")) or 0
    ltd = to_float(getv(ws, row, hmap, "장기부채")) or 0
    std = to_float(getv(ws, row, hmap, "단기위험부채")) or 0

    if "주당순현금(린치식)" in hmap:
        setv(ws, row, hmap, "주당순현금(린치식)", (cash + sec - ltd) / latest_shares)
    if "주당순현금(보수형)" in hmap:
        setv(ws, row, hmap, "주당순현금(보수형)", (cash + sec - ltd - std) / latest_shares)

    # EPS / DPS / 주당FCF는 총액이 없으므로 factor 조정
    for col in ["EPS(FY2025)", "현금배당금(FY2025 DPS)", "주당잉여현금흐름"]:
        v = to_float(getv(ws, row, hmap, col))
        if v is not None and factor not in (None, 0):
            setv(ws, row, hmap, col, v / factor)


def update_timing_sheet_prices(wb, main_ws, main_hmap):
    if "타이밍자동체크" not in wb.sheetnames:
        return

    t_ws = wb["타이밍자동체크"]
    t_header_row = find_header_row(t_ws, PRICE_REQUIRED, max_rows=10)
    if not t_header_row:
        return
    t_hmap = header_map(t_ws, t_header_row)

    price_by_code = {}
    for r in range(main_ws.max_row, t_header_row, -1):
        code = normalize_code(getv(main_ws, r, main_hmap, "종목코드"))
        if not code:
            continue
        price = to_float(getv(main_ws, r, main_hmap, "현재가"))
        if price is not None:
            price_by_code[code] = price

    for r in range(t_header_row + 1, t_ws.max_row + 1):
        code = normalize_code(getv(t_ws, r, t_hmap, "종목코드"))
        if not code:
            continue
        price = price_by_code.get(code)
        if price is not None and "현재가" in t_hmap:
            setv(t_ws, r, t_hmap, "현재가", price)


def add_note_sheet(wb, note_lines):
    if "업데이트메모" in wb.sheetnames:
        del wb["업데이트메모"]
    ws = wb.create_sheet("업데이트메모")
    for i, line in enumerate(note_lines, start=1):
        ws.cell(i, 1).value = line
    ws.column_dimensions["A"].width = 120


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="input_xlsx", required=True)
    ap.add_argument("--out", dest="output_xlsx", required=True)
    ap.add_argument("--asof", default=None, help="YYYY-MM-DD. 기본은 오늘")
    ap.add_argument("--main-sheet", default=None, help="메인 시트명 직접 지정")
    ap.add_argument("--only-code", default=None, help="특정 종목코드만 테스트")
    args = ap.parse_args()

    in_path = Path(args.input_xlsx)
    out_path = Path(args.output_xlsx)

    wb = openpyxl.load_workbook(in_path)

    # main sheet 탐색
    if args.main_sheet:
        if args.main_sheet not in wb.sheetnames:
            raise SystemExit(f"메인 시트를 찾지 못함: {args.main_sheet}")
        main_ws = wb[args.main_sheet]
    else:
        main_ws = None
        for name in wb.sheetnames:
            ws = wb[name]
            hdr = find_header_row(ws, MAIN_REQUIRED, max_rows=10)
            if hdr:
                main_ws = ws
                break
        if main_ws is None:
            raise SystemExit("메인 시트를 자동 탐지하지 못함")

    header_row = find_header_row(main_ws, MAIN_REQUIRED, max_rows=10)
    if not header_row:
        raise SystemExit("메인 시트 헤더 행 탐지 실패")
    hmap = header_map(main_ws, header_row)

    updated = 0
    split_adjusted = 0
    errors = []

    for r in range(header_row + 1, main_ws.max_row + 1):
        code = normalize_code(getv(main_ws, r, hmap, "종목코드"))
        name = getv(main_ws, r, hmap, "종목명")
        if not code:
            continue
        if args.only_code and code != normalize_code(args.only_code):
            continue

        try:
            latest_date, latest_close, latest_shares = latest_ohlcv_and_shares(code, asof=args.asof)
            old_shares = to_float(getv(main_ws, r, hmap, "발행주식수"))

            setv(main_ws, r, hmap, "현재가", latest_close)
            normalize_per_share_values(main_ws, r, hmap, latest_shares)
            recalc_graham_values(main_ws, r, hmap)
            recalc_price_dependent(main_ws, r, hmap)

            updated += 1
            if old_shares not in (None, 0) and int(old_shares) != int(latest_shares):
                split_adjusted += 1
        except Exception as e:
            errors.append(f"{code} {name}: {e}")

    update_timing_sheet_prices(wb, main_ws, hmap)

    note_lines = [
        "이 파일은 최신 거래일 KRX 종가와 상장주식수로 보정됨.",
        f"보정 기준일: {args.asof or date.today().isoformat()}",
        f"업데이트 종목 수: {updated}",
        f"상장주식수 변경 감지(액면분할/병합 가능성) 종목 수: {split_adjusted}",
        "",
        "핵심 보정 로직:",
        "1) 현재가 <- 최신 거래일 종가",
        "2) 발행주식수 <- 최신 상장주식수",
        "3) EPS/DPS/주당FCF <- 상장주식수 비율로 정규화",
        "4) 주당순현금(린치/보수형) <- absolute totals / 최신 상장주식수로 재계산",
        "5) 그레이엄 내재가치/괴리율, 순현금차감PER, 배당수익률, 린치PER배수, 종합판정 재계산",
        "",
        "주의:",
        "- pykrx는 장중 실시간 틱이 아니라 최신 거래일 기준 데이터로 쓰는 게 안전함.",
        "- 장중 실시간 매수판단은 별도 '현재가_실시간' 컬럼을 두고 overlay 하는 방식을 권장.",
        "",
    ]
    if errors:
        note_lines.append("오류 종목:")
        note_lines.extend(errors[:200])

    add_note_sheet(wb, note_lines)
    wb.save(out_path)
    print(f"saved: {out_path}")
    print(f"updated={updated}, split_adjusted={split_adjusted}, errors={len(errors)}")


if __name__ == "__main__":
    main()
