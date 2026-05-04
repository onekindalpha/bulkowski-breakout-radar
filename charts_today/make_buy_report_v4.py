#!/usr/bin/env python3
"""
make_buy_report_v4.py  (BUY_NOW first + colored Excel)

What you asked for:
- BUY_NOW / BUY_NOW_REASON are ALWAYS the first columns (so you can't miss them).
- Also writes an XLSX with colors:
    BUY_NOW=Y  -> green row
    BUY_NOW=N  -> light gray row
- Still writes CSV.
- Prints a concise terminal summary:
    BUY_NOW=Y list (or none)
    NEAR_BUY list (CloseBreak=Y & VolConfirm=Y but failed BUY_NOW)

BUY_NOW rule (strict by default, tuneable):
BUY_NOW = all true:
  1) breakout_confirmed_close == True
  2) breakout_volume_confirmed == True
  3) breakout_confirmed_intraday == True   (required by default if column exists; you typically run --intraday)
  4) price >= daily_break_level
  5) price <= daily_break_level*(1+buy_chase_pct/100)   (default 2%)
  6) room_to_weekly_r1_pct >= buy_min_room_pct          (default 1%)
  7) buy_rsi_min <= rsi14 <= buy_rsi_max                (default 45..75)

Usage:
  python make_buy_report_v4.py --top 15
  python make_buy_report_v4.py --top 15 --xlsx
  python make_buy_report_v4.py --tickers XLE,XOP,... --xlsx
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd

KST = ZoneInfo("Asia/Seoul")


def newest_report_path() -> Path:
    stamped = sorted(Path(".").glob("report_v2_*_KST.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
    if stamped:
        return stamped[0]
    p = Path("report_v2.csv")
    if not p.exists():
        raise FileNotFoundError("No report found: report_v2.csv or report_v2_*_KST.csv")
    return p


def _fmt_bool(x) -> str:
    try:
        return "Y" if bool(x) else "N"
    except Exception:
        return "N"


def _to_float(x):
    try:
        return float(x)
    except Exception:
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--top", type=int, default=15)
    ap.add_argument("--tickers", default="")
    ap.add_argument("--xlsx", action="store_true", help="write a colored Excel file too")
    # BUY_NOW knobs
    ap.add_argument("--buy-chase-pct", type=float, default=2.0, help="max % above break_level allowed (default 2.0)")
    ap.add_argument("--buy-min-room-pct", type=float, default=1.0, help="min room_to_weekly_r1_pct (default 1.0)")
    ap.add_argument("--buy-rsi-min", type=float, default=45.0, help="min RSI14 (default 45)")
    ap.add_argument("--buy-rsi-max", type=float, default=75.0, help="max RSI14 (default 75)")
    ap.add_argument("--require-intraday", action="store_true", help="require intraday hold even if column missing")
    ap.add_argument("--no-intraday", action="store_true", help="do NOT require intraday hold even if present")
    args = ap.parse_args()

    rp = newest_report_path()
    df = pd.read_csv(rp, comment="#")
    df.columns = [c.strip() for c in df.columns]

    want_cols = [
        "ticker","grade","score","price","gap_pct","rsi14",
        "daily_break_level","daily_breakout","daily_retest",
        "breakout_confirmed_close","breakout_volume_confirmed","breakout_confirmed_intraday",
        "weekly_r1","room_to_weekly_r1_pct","px_vs_sma50","px_vs_sma200",
    ]
    cols = [c for c in want_cols if c in df.columns]
    out = df[cols].copy()

    if args.tickers.strip():
        keep = {t.strip().upper() for t in args.tickers.split(",") if t.strip()}
        out = out[out["ticker"].astype(str).str.upper().isin(keep)].copy()
    else:
        out = out.head(max(1, args.top)).copy()

    # summary reason
    def build_reason(r):
        parts = []
        if "grade" in out.columns and pd.notna(r.get("grade")):
            parts.append(f"Grade={r.get('grade')}")
        if "score" in out.columns and pd.notna(r.get("score")):
            parts.append(f"Score={r.get('score')}")
        if "breakout_confirmed_intraday" in out.columns:
            parts.append(f"IntradayHold={_fmt_bool(r.get('breakout_confirmed_intraday'))}")
        if "breakout_confirmed_close" in out.columns:
            parts.append(f"CloseBreak={_fmt_bool(r.get('breakout_confirmed_close'))}")
        if "breakout_volume_confirmed" in out.columns:
            parts.append(f"VolConfirm={_fmt_bool(r.get('breakout_volume_confirmed'))}")
        if "daily_retest" in out.columns:
            parts.append(f"Retest={_fmt_bool(r.get('daily_retest'))}")
        if "rsi14" in out.columns and pd.notna(r.get("rsi14")):
            parts.append(f"RSI={float(r.get('rsi14')):.1f}")
        if "room_to_weekly_r1_pct" in out.columns and pd.notna(r.get("room_to_weekly_r1_pct")):
            parts.append(f"RoomR1={float(r.get('room_to_weekly_r1_pct')):.1f}%")
        if "gap_pct" in out.columns and pd.notna(r.get("gap_pct")):
            parts.append(f"Gap={float(r.get('gap_pct')):.2f}%")
        if "daily_break_level" in out.columns and pd.notna(r.get("daily_break_level")):
            parts.append(f"BreakLvl={float(r.get('daily_break_level')):.2f}")
        return " | ".join(parts)

    out["reason"] = out.apply(build_reason, axis=1)

    # BUY_NOW decision
    has_intraday_col = "breakout_confirmed_intraday" in out.columns
    if args.no_intraday:
        require_intraday = False
    elif args.require_intraday:
        require_intraday = True
    else:
        # default: require intraday if the column exists
        require_intraday = has_intraday_col

    def buy_eval(r):
        fail = []

        # 1) close confirm
        if "breakout_confirmed_close" in out.columns:
            if not bool(r.get("breakout_confirmed_close")):
                fail.append("CloseBreak=N")
        else:
            fail.append("CloseBreak=missing")

        # 2) volume confirm
        if "breakout_volume_confirmed" in out.columns:
            if not bool(r.get("breakout_volume_confirmed")):
                fail.append("VolConfirm=N")
        else:
            fail.append("VolConfirm=missing")

        # 3) intraday hold
        if require_intraday:
            if not has_intraday_col:
                fail.append("IntradayHold=missing")
            else:
                if not bool(r.get("breakout_confirmed_intraday")):
                    fail.append("IntradayHold=N")

        price = _to_float(r.get("price"))
        lvl = _to_float(r.get("daily_break_level"))
        rsi = _to_float(r.get("rsi14"))
        room = _to_float(r.get("room_to_weekly_r1_pct"))

        # 4) price >= level
        if price is None or lvl is None:
            fail.append("Price/Level=missing")
        else:
            if price < lvl:
                fail.append("BelowBreakLvl")
            if price > lvl * (1.0 + args.buy_chase_pct/100.0):
                fail.append(f"Chase>{args.buy_chase_pct:.1f}%")

        # 6) room
        if room is None:
            fail.append("RoomR1=missing")
        else:
            if room < args.buy_min_room_pct:
                fail.append(f"RoomR1<{args.buy_min_room_pct:.1f}%")

        # 7) RSI band
        if rsi is None:
            fail.append("RSI=missing")
        else:
            if rsi < args.buy_rsi_min:
                fail.append(f"RSI<{args.buy_rsi_min:.0f}")
            if rsi > args.buy_rsi_max:
                fail.append(f"RSI>{args.buy_rsi_max:.0f}")

        buy_now = (len(fail) == 0)
        return ("Y" if buy_now else "N"), ("OK" if buy_now else ";".join(fail))

    out[["BUY_NOW","BUY_NOW_REASON"]] = out.apply(lambda r: pd.Series(buy_eval(r)), axis=1)

    # reorder columns: BUY_NOW first
    front = ["BUY_NOW","BUY_NOW_REASON","ticker","price","daily_break_level"]
    new_cols = []
    for c in front:
        if c in out.columns and c not in new_cols:
            new_cols.append(c)
    for c in out.columns:
        if c not in new_cols:
            new_cols.append(c)
    out = out[new_cols].copy()

    ts = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
    csv_name = f"buy_report_{ts}_KST.csv"
    out.to_csv(csv_name, index=False)

    buy_cnt = int((out["BUY_NOW"] == "Y").sum())
    print(f"Source: {rp.name}")
    print(f"Saved: {csv_name} (rows={len(out)} | BUY_NOW=Y: {buy_cnt})")
    print(
        "BUY_NOW rule: CloseBreak=Y & VolConfirm=Y"
        + (" & IntradayHold=Y" if require_intraday else "")
        + f" & price in [BreakLvl, BreakLvl*(1+{args.buy_chase_pct:.1f}%)]"
        + f" & RoomR1>={args.buy_min_room_pct:.1f}%"
        + f" & RSI in [{args.buy_rsi_min:.0f},{args.buy_rsi_max:.0f}]"
    )

    print("\n=== BUY_NOW (Y) ===")
    if buy_cnt == 0:
        print("(none)")
    else:
        show = [c for c in ["ticker","price","daily_break_level","BUY_NOW_REASON"] if c in out.columns]
        print(out[out["BUY_NOW"]=="Y"][show].to_string(index=False))

    # Near buys: close+volume OK, but failed BUY_NOW
    near = out.copy()
    if "breakout_confirmed_close" in near.columns:
        near = near[near["breakout_confirmed_close"].astype(str).isin(["True","true","1","Y","y"]) | (near["breakout_confirmed_close"]==True)]
    if "breakout_volume_confirmed" in near.columns:
        near = near[near["breakout_volume_confirmed"].astype(str).isin(["True","true","1","Y","y"]) | (near["breakout_volume_confirmed"]==True)]
    near = near[near["BUY_NOW"]=="N"]
    print("\n=== NEAR_BUY (CloseBreak=Y & VolConfirm=Y but failed BUY_NOW) ===")
    if near.empty:
        print("(none)")
    else:
        show = [c for c in ["ticker","price","daily_break_level","breakout_confirmed_intraday","room_to_weekly_r1_pct","rsi14","BUY_NOW_REASON"] if c in near.columns]
        print(near[show].head(15).to_string(index=False))

    if args.xlsx:
        try:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.formatting.rule import FormulaRule

            xlsx_name = f"buy_report_{ts}_KST.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "buy_report"

            # write rows
            for r in dataframe_to_rows(out, index=False, header=True):
                ws.append(r)

            # freeze header
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            # find BUY_NOW column index
            header = [c.value for c in ws[1]]
            buy_col = header.index("BUY_NOW") + 1 if "BUY_NOW" in header else 1

            # styles
            fill_y = PatternFill("solid", fgColor="C6EFCE")  # light green
            fill_n = PatternFill("solid", fgColor="F2F2F2")  # light gray
            font_y = Font(bold=True, color="006100")
            align = Alignment(vertical="center")

            # apply conditional formatting to entire row range based on BUY_NOW cell
            last_row = ws.max_row
            last_col = ws.max_column
            # Excel row starts at 2 for data
            # formula uses absolute col for BUY_NOW and relative row
            col_letter = ws.cell(row=1, column=buy_col).column_letter

            # Y rows
            ws.conditional_formatting.add(
                f"A2:{ws.cell(row=last_row, column=last_col).coordinate}",
                FormulaRule(formula=[f'${col_letter}2="Y"'], fill=fill_y, font=font_y)
            )
            # N rows (light gray)
            ws.conditional_formatting.add(
                f"A2:{ws.cell(row=last_row, column=last_col).coordinate}",
                FormulaRule(formula=[f'${col_letter}2="N"'], fill=fill_n)
            )

            # header style
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align

            # alignment for all
            for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
                for cell in row:
                    cell.alignment = align

            # basic column widths
            for j, name in enumerate(header, 1):
                width = max(10, min(35, len(str(name)) + 2))
                ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = width

            wb.save(xlsx_name)
            print(f"\nSaved: {xlsx_name} (colored)")
        except Exception as e:
            print(f"\n[warn] XLSX not created: {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
