#!/usr/bin/env python3
"""
make_buy_report_v5.py  (BUY_NOW first + colored Excel + PRIORITY ranking)

What this adds vs v4:
- PRIORITY_SCORE + PRIORITY_RANK (1 = best)
- breakout_pct_vs_level (% above/below break line)
- chase_flag (CHASE if breakout_pct_vs_level > buy_chase_pct)
- near_buy flag (CloseBreak=Y & VolConfirm=Y but BUY_NOW=N)

Output:
- buy_report_<stamp>_KST.csv
- buy_report_<stamp>_KST.xlsx  (if --xlsx)

BUY_NOW rule (strict, same intent as before):
BUY_NOW=Y iff ALL:
  CloseBreak=Y
  VolConfirm=Y
  IntradayHold=Y   (required by default if column exists; since you run --intraday)
  price >= BreakLvl
  price <= BreakLvl*(1+buy_chase_pct/100)   (default 2%)
  room_to_weekly_r1_pct >= buy_min_room_pct (default 1%)
  RSI in [buy_rsi_min, buy_rsi_max]         (default 45..75)

PRIORITY_SCORE (0..100-ish, transparent, no magic):
+ 20  if CloseBreak=Y
+ 20  if VolConfirm=Y
+ 15  if IntradayHold=Y
+ 10  if Retest=Y
+ Room score: up to +20 (room_to_weekly_r1_pct, capped at 10% -> 20 points)
+ Breakout distance score: up to +20 (best when 0~+1% above break; penalize chasing)
+ RSI score: up to +10 (best near RSI=60; degrades as you move away)
+ Trend score: up to +5  (px_vs_sma50, capped)

Then rows are sorted:
  BUY_NOW desc, PRIORITY_SCORE desc, score desc

Usage:
  python make_buy_report_v5.py --top 20 --xlsx
  python make_buy_report_v5.py --tickers XLE,XOP,... --xlsx
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd

KST = ZoneInfo("Asia/Seoul")


def newest_report_path() -> Path:
    stamped = sorted(Path(".").glob("report_v2_*_KST.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
    if stamped:
        return stamped[0]
    p = Path("report_v2.csv")
    if not p.exists():
        raise FileNotFoundError("No report found: report_v2.csv or report_v2_*_KST.csv")
    return p


def _fmt_bool(x) -> str:
    try:
        return "Y" if bool(x) else "N"
    except Exception:
        return "N"


def _to_float(x):
    try:
        return float(x)
    except Exception:
        return None


def _clamp(x, lo, hi):
    return max(lo, min(hi, x))


def _ensure_ticker_column(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize a ticker/symbol column to 'ticker'.

    Handles common cases:
      - column named Symbol/Ticker (case-insensitive)
      - first unnamed index column from index=True CSV save
      - ticker stored as index
    """
    df.columns = [str(c).strip() for c in df.columns]

    # Case-insensitive rename
    lower_map = {str(c).strip().lower(): c for c in df.columns}
    if "ticker" in lower_map and lower_map["ticker"] != "ticker":
        return df.rename(columns={lower_map["ticker"]: "ticker"})
    if "ticker" in lower_map:
        return df
    if "symbol" in lower_map:
        return df.rename(columns={lower_map["symbol"]: "ticker"})

    # 'Unnamed: 0' artifact (saved with index=True)
    for c in list(df.columns):
        if str(c).strip().lower().startswith("unnamed"):
            s = df[c].astype(str)
            if (s.str.len().median() <= 8) and (s.str.match(r"^[A-Za-z0-9=\-\.\^]+$").mean() > 0.8):
                return df.rename(columns={c: "ticker"})

    # Ticker in index
    if df.index.name and str(df.index.name).strip().lower() in {"ticker", "symbol"}:
        return df.reset_index().rename(columns={df.index.name: "ticker"})

    # Heuristic: first column looks like ticker
    if len(df.columns) >= 1:
        c0 = df.columns[0]
        s = df[c0].astype(str)
        if (s.str.len().median() <= 8) and (s.str.match(r"^[A-Za-z0-9=\-\.\^]+$").mean() > 0.8):
            return df.rename(columns={c0: "ticker"})

    return df




_TWO_X_TICKERS = {
    # 2x / leveraged ETFs commonly treated with higher extension tolerance
    "ERX","DIG","UYM","UCO","BOIL","KOLD","GUSH","DRIP","NUGT","DUST","LABU","LABD","SOXL","SOXS","TQQQ","SQQQ",
}

def _as_bool(x) -> bool:
    if isinstance(x, bool):
        return x
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return False
    s = str(x).strip().lower()
    return s in {"1","true","t","y","yes"}

def _bdays_since(date_like, today_date) -> int | None:
    dt = pd.to_datetime(date_like, errors="coerce")
    if pd.isna(dt):
        return None
    bd = dt.date()
    # trading-day-ish: weekday business days, ignore holidays
    rng = pd.bdate_range(start=bd, end=today_date)
    return max(0, len(rng) - 1)

def _print_legacy_breakout_block(df_full: pd.DataFrame):
    needed = ["ticker","price","break_level","breakout_date","vol_confirmed","hold_confirmed","rsi14"]
    missing = [c for c in needed if c not in df_full.columns]
    print("\n=== LEGACY_BREAKOUT_BUY ===")
    if missing:
        print(f"(skipped: missing columns: {', '.join(missing)})")
        return

    today = datetime.now(KST).date()

    rows = []
    for _, r in df_full.iterrows():
        t = str(r.get("ticker", "")).strip().upper()
        lvl = _to_float(r.get("break_level"))
        px = _to_float(r.get("price"))
        if not t or lvl is None or px is None or lvl <= 0:
            continue
        if px <= lvl:
            continue  # legacy focuses on "확실히 넘은 애들"

        vol = _as_bool(r.get("vol_confirmed"))
        hold = _as_bool(r.get("hold_confirmed"))
        rsi = _to_float(r.get("rsi14"))
        bd = r.get("breakout_date")
        age = _bdays_since(bd, today)
        if age is None:
            age = 999

        diff = px - lvl
        ext_pct = (diff / lvl) * 100.0

        ext_limit = 2.5 if t in _TWO_X_TICKERS else 1.5

        bucket = "WATCH"
        if vol and hold:
            if (age <= 5) and (ext_pct <= ext_limit) and (rsi is None or rsi <= 72):
                bucket = "BUY_CANDIDATE"
            elif (ext_pct >= 4.5) or (rsi is not None and rsi > 75):
                bucket = "LATE"
            else:
                bucket = "WATCH"
        else:
            # breakout above line but confirmations missing -> WATCH (riskier)
            bucket = "WATCH"

        rows.append({
            "bucket": bucket,
            "ticker": t,
            "price": px,
            "break_level": lvl,
            "diff": diff,
            "ext_pct": ext_pct,
            "breakout_date": str(bd),
            "age": age,
            "vol": vol,
            "hold": hold,
            "rsi": rsi,
        })

    if not rows:
        print("(none)")
        return

    d = pd.DataFrame(rows)

    order = {"BUY_CANDIDATE": 0, "WATCH": 1, "LATE": 2}
    d["bucket_rank"] = d["bucket"].map(order).fillna(9).astype(int)
    d = d.sort_values(["bucket_rank","age","ext_pct"], ascending=[True, True, True])

    def fmt_line(rr):
        t = rr["ticker"]
        px = rr["price"]; lvl = rr["break_level"]
        diff = rr["diff"]; ext = rr["ext_pct"]
        bd = rr["breakout_date"]; age = int(rr["age"]) if rr["age"] != 999 else "?"
        line = f"{t:<6} {px:>8.2f} > {lvl:>8.2f}  (+{diff:.2f}, +{ext:.2f}%)  breakout_date={bd}  hold={age}d"
        if not rr["vol"] or not rr["hold"]:
            line += f"  (vol={_fmt_bool(rr['vol'])}, hold={_fmt_bool(rr['hold'])})"
        return line

    for bucket in ["BUY_CANDIDATE", "WATCH", "LATE"]:
        sub = d[d["bucket"] == bucket]
        print(f"\n{bucket}")
        if sub.empty:
            print("(none)")
            continue
        for _, rr in sub.iterrows():
            print(fmt_line(rr))

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--top", type=int, default=20, help="how many rows to output after ranking (default 20)")
    ap.add_argument("--tickers", default="", help="comma tickers; overrides --top")
    ap.add_argument("--xlsx", action="store_true", help="write a colored Excel file too")
    # BUY_NOW knobs
    ap.add_argument("--buy-chase-pct", type=float, default=2.0)
    ap.add_argument("--buy-min-room-pct", type=float, default=1.0)
    ap.add_argument("--buy-rsi-min", type=float, default=45.0)
    ap.add_argument("--buy-rsi-max", type=float, default=75.0)
    ap.add_argument("--require-intraday", action="store_true")
    ap.add_argument("--no-intraday", action="store_true")
    args = ap.parse_args()

    rp = newest_report_path()
    df = pd.read_csv(rp, comment="#")
    df.columns = [c.strip() for c in df.columns]
    df = _ensure_ticker_column(df)
    df_full = df.copy()

    # We rank from the FULL report, then take top N
    want_cols = [
        "ticker","grade","score","price","gap_pct","rsi14",
        "daily_break_level","daily_breakout","daily_retest",
        "breakout_confirmed_close","breakout_volume_confirmed","breakout_confirmed_intraday",
        "weekly_r1","room_to_weekly_r1_pct","px_vs_sma50","px_vs_sma200",
    ]
    cols = [c for c in want_cols if c in df.columns]
    out = df[cols].copy()

    # Optional ticker filter (exact)
    if args.tickers.strip():
        keep = {t.strip().upper() for t in args.tickers.split(",") if t.strip()}
        out = out[out["ticker"].astype(str).str.upper().isin(keep)].copy()

    # human readable reason (quick scan)
    def build_reason(r):
        parts = []
        if "grade" in out.columns and pd.notna(r.get("grade")):
            parts.append(f"Grade={r.get('grade')}")
        if "score" in out.columns and pd.notna(r.get("score")):
            parts.append(f"Score={r.get('score')}")
        if "breakout_confirmed_intraday" in out.columns:
            parts.append(f"IntradayHold={_fmt_bool(r.get('breakout_confirmed_intraday'))}")
        if "breakout_confirmed_close" in out.columns:
            parts.append(f"CloseBreak={_fmt_bool(r.get('breakout_confirmed_close'))}")
        if "breakout_volume_confirmed" in out.columns:
            parts.append(f"VolConfirm={_fmt_bool(r.get('breakout_volume_confirmed'))}")
        if "daily_retest" in out.columns:
            parts.append(f"Retest={_fmt_bool(r.get('daily_retest'))}")
        if "rsi14" in out.columns and pd.notna(r.get("rsi14")):
            parts.append(f"RSI={float(r.get('rsi14')):.1f}")
        if "room_to_weekly_r1_pct" in out.columns and pd.notna(r.get("room_to_weekly_r1_pct")):
            parts.append(f"RoomR1={float(r.get('room_to_weekly_r1_pct')):.1f}%")
        if "gap_pct" in out.columns and pd.notna(r.get("gap_pct")):
            parts.append(f"Gap={float(r.get('gap_pct')):.2f}%")
        if "daily_break_level" in out.columns and pd.notna(r.get("daily_break_level")):
            parts.append(f"BreakLvl={float(r.get('daily_break_level')):.2f}")
        return " | ".join(parts)

    out["reason"] = out.apply(build_reason, axis=1)

    # Decide whether intraday is required for BUY_NOW
    has_intraday_col = "breakout_confirmed_intraday" in out.columns
    if args.no_intraday:
        require_intraday = False
    elif args.require_intraday:
        require_intraday = True
    else:
        require_intraday = has_intraday_col  # default: require if available

    # Compute breakout_pct_vs_level + chase_flag
    def breakout_pct(r):
        price = _to_float(r.get("price"))
        lvl = _to_float(r.get("daily_break_level"))
        if price is None or lvl is None or lvl == 0:
            return None
        return (price / lvl - 1.0) * 100.0

    out["breakout_pct_vs_level"] = out.apply(breakout_pct, axis=1)
    out["chase_flag"] = out["breakout_pct_vs_level"].apply(
        lambda x: "CHASE" if (x is not None and pd.notna(x) and x > args.buy_chase_pct) else ""
    )

    # BUY_NOW
    def buy_eval(r):
        fail = []

        if "breakout_confirmed_close" in out.columns:
            if not bool(r.get("breakout_confirmed_close")):
                fail.append("CloseBreak=N")
        else:
            fail.append("CloseBreak=missing")

        if "breakout_volume_confirmed" in out.columns:
            if not bool(r.get("breakout_volume_confirmed")):
                fail.append("VolConfirm=N")
        else:
            fail.append("VolConfirm=missing")

        if require_intraday:
            if not has_intraday_col:
                fail.append("IntradayHold=missing")
            else:
                if not bool(r.get("breakout_confirmed_intraday")):
                    fail.append("IntradayHold=N")

        price = _to_float(r.get("price"))
        lvl = _to_float(r.get("daily_break_level"))
        rsi = _to_float(r.get("rsi14"))
        room = _to_float(r.get("room_to_weekly_r1_pct"))

        if price is None or lvl is None:
            fail.append("Price/Level=missing")
        else:
            if price < lvl:
                fail.append("BelowBreakLvl")
            if price > lvl * (1.0 + args.buy_chase_pct/100.0):
                fail.append(f"Chase>{args.buy_chase_pct:.1f}%")

        if room is None:
            fail.append("RoomR1=missing")
        else:
            if room < args.buy_min_room_pct:
                fail.append(f"RoomR1<{args.buy_min_room_pct:.1f}%")

        if rsi is None:
            fail.append("RSI=missing")
        else:
            if rsi < args.buy_rsi_min:
                fail.append(f"RSI<{args.buy_rsi_min:.0f}")
            if rsi > args.buy_rsi_max:
                fail.append(f"RSI>{args.buy_rsi_max:.0f}")

        buy_now = (len(fail) == 0)
        return ("Y" if buy_now else "N"), ("OK" if buy_now else ";".join(fail))

    out[["BUY_NOW","BUY_NOW_REASON"]] = out.apply(lambda r: pd.Series(buy_eval(r)), axis=1)

    # near_buy flag
    def is_near(r):
        return (bool(r.get("breakout_confirmed_close")) and bool(r.get("breakout_volume_confirmed")) and r.get("BUY_NOW") == "N")
    out["near_buy"] = out.apply(is_near, axis=1)

    # PRIORITY_SCORE
    def priority_score(r):
        score = 0.0
        # flags
        if bool(r.get("breakout_confirmed_close")): score += 20
        if bool(r.get("breakout_volume_confirmed")): score += 20
        if "breakout_confirmed_intraday" in out.columns and bool(r.get("breakout_confirmed_intraday")): score += 15
        if bool(r.get("daily_retest")): score += 10

        # room score (cap at 10% => 20 points)
        room = _to_float(r.get("room_to_weekly_r1_pct"))
        if room is not None:
            score += _clamp(room, 0, 10) * 2.0

        # breakout distance score: best 0..+1%, penalize >2%
        pct = _to_float(r.get("breakout_pct_vs_level"))
        if pct is None:
            pass
        else:
            if pct < 0:
                score += -5
            elif pct <= 1:
                score += 20 - pct*5
            elif pct <= 2:
                score += 15 - (pct-1)*10
            else:
                score += 5 - _clamp((pct-2)*5, 0, 20)

        # RSI score (best near 60): 0..10
        rsi = _to_float(r.get("rsi14"))
        if rsi is not None:
            score += _clamp(10 - abs(rsi-60)/3, 0, 10)

        # trend score from px_vs_sma50 (cap at 10% => 5 points)
        sma50 = _to_float(r.get("px_vs_sma50"))
        if sma50 is not None:
            score += _clamp(sma50, -5, 10) * 0.5  # -2.5..+5

        return round(score, 2)

    out["PRIORITY_SCORE"] = out.apply(priority_score, axis=1)

    # Sort + take top
    out["BUY_NOW_sort"] = out["BUY_NOW"].map({"Y":1,"N":0}).fillna(0).astype(int)
    out = out.sort_values(["BUY_NOW_sort","PRIORITY_SCORE","score"], ascending=[False,False,False]).drop(columns=["BUY_NOW_sort"])
    if not args.tickers.strip():
        out = out.head(max(1, args.top)).copy()

    # rank number
    out["PRIORITY_RANK"] = range(1, len(out)+1)

    # reorder columns (BUY_NOW first)
    front = ["BUY_NOW","BUY_NOW_REASON","PRIORITY_RANK","PRIORITY_SCORE","near_buy","chase_flag",
             "ticker","price","daily_break_level","breakout_pct_vs_level","reason"]
    new_cols = []
    for c in front:
        if c in out.columns and c not in new_cols:
            new_cols.append(c)
    for c in out.columns:
        if c not in new_cols:
            new_cols.append(c)
    out = out[new_cols].copy()

    ts = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
    csv_name = f"buy_report_{ts}_KST.csv"
    out.to_csv(csv_name, index=False)

    buy_cnt = int((out["BUY_NOW"] == "Y").sum())
    print(f"Source: {rp.name}")
    print(f"Saved: {csv_name} (rows={len(out)} | BUY_NOW=Y: {buy_cnt})")
    print(
        "BUY_NOW rule: CloseBreak=Y & VolConfirm=Y"
        + (" & IntradayHold=Y" if require_intraday else "")
        + f" & price in [BreakLvl, BreakLvl*(1+{args.buy_chase_pct:.1f}%)]"
        + f" & RoomR1>={args.buy_min_room_pct:.1f}%"
        + f" & RSI in [{args.buy_rsi_min:.0f},{args.buy_rsi_max:.0f}]"
    )

    _print_legacy_breakout_block(df_full)

    print("\n=== BUY_NOW (Y) ===")
    cols_buy = [c for c in ["ticker","price","daily_break_level","breakout_pct_vs_level","BUY_NOW_REASON","PRIORITY_SCORE"] if c in out.columns]
    buy_rows = out.loc[out["BUY_NOW"]=="Y", cols_buy]
    if buy_rows.empty:
        print("(none)")
    else:
        print(buy_rows.to_string(index=False))

    print("\n=== TOP PRIORITY (ranked) ===")
    show = [c for c in ["PRIORITY_RANK","ticker","BUY_NOW","PRIORITY_SCORE","breakout_pct_vs_level","room_to_weekly_r1_pct","rsi14","chase_flag","near_buy","BUY_NOW_REASON"] if c in out.columns]
    print(out[show].head(15).to_string(index=False))

    if args.xlsx:
        try:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.formatting.rule import FormulaRule

            xlsx_name = f"buy_report_{ts}_KST.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "buy_report"

            for r in dataframe_to_rows(out, index=False, header=True):
                ws.append(r)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            header = [c.value for c in ws[1]]
            buy_col = header.index("BUY_NOW") + 1 if "BUY_NOW" in header else 1
            near_col = header.index("near_buy") + 1 if "near_buy" in header else None
            chase_col = header.index("chase_flag") + 1 if "chase_flag" in header else None

            fill_buy = PatternFill("solid", fgColor="C6EFCE")  # green
            fill_near = PatternFill("solid", fgColor="FFF2CC") # yellow
            fill_chase = PatternFill("solid", fgColor="F8CBAD")# red-ish
            fill_n = PatternFill("solid", fgColor="F2F2F2")    # gray
            font_buy = Font(bold=True, color="006100")
            align = Alignment(vertical="center")

            last_row = ws.max_row
            last_col = ws.max_column
            rng = f"A2:{ws.cell(row=last_row, column=last_col).coordinate}"
            buy_letter = ws.cell(row=1, column=buy_col).column_letter

            # BUY rows (green)
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f'${buy_letter}2="Y"'], fill=fill_buy, font=font_buy)
            )
            # NEAR rows (yellow) but only if BUY_NOW is N
            if near_col is not None:
                near_letter = ws.cell(row=1, column=near_col).column_letter
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(formula=[f'AND(${buy_letter}2="N",${near_letter}2=TRUE)'], fill=fill_near)
                )
            # CHASE rows (red) but only if BUY_NOW is N (warning)
            if chase_col is not None:
                chase_letter = ws.cell(row=1, column=chase_col).column_letter
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(formula=[f'AND(${buy_letter}2="N",${chase_letter}2="CHASE")'], fill=fill_chase)
                )
            # Remaining N rows (gray) - applied last (lowest priority), might be overridden by above rules
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[f'${buy_letter}2="N"'], fill=fill_n)
            )

            # header style
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align

            for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=last_col):
                for cell in row:
                    cell.alignment = align

            for j, name in enumerate(header, 1):
                width = max(10, min(40, len(str(name)) + 2))
                ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = width

            wb.save(xlsx_name)
            print(f"\nSaved: {xlsx_name} (colored)")
        except Exception as e:
            print(f"\n[warn] XLSX not created: {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()
